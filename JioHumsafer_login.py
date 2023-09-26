from selenium.common.exceptions import WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from selenium import webdriver
import configparser
import datetime
import openpyxl
import requests
import time
import csv
import pandas as pd

'''# Read the Excel file
Get_input=str(input("Please enter xl file to fetch list items (in .xlsx naming format): "))
df = pd.read_excel(Get_input)
# Extract data from row 2
row_number = 1
row_data = df.iloc[row_number]
# Convert the row data to a Python list
list = row_data.tolist()'''

list=['Engine RPM', ' Acc Pedal Pos(%) ', ' Parking Brake Switch ', ' Speed(kmph) ', ' Brake Switch ',
      ' Clutch Switch ', 'Coolant Temperature(C)', 'Eng oil Temperature(C)','Fuel used','Engine hours','Eng Oil Pressure','Distance travelled','Gear','Odo','Air pressure 1','Air pressure 2','H2 leak sensor 1','H2 leak sensor 2','H2 High pressure','Tank Temperature 1','Tank temperature 2','H2 Storage Lvl (%)','H2 Storage Lvl (Kg)','DTE','Avg Fuel Economy','longitude','latitude','GPS Speed','Accuracy (m)','Altitude']

def jiohumsufer_login():
    driver = webdriver.Chrome()
    wait = WebDriverWait(driver, 30)
    with open("config.ini", "r") as file:
        config = configparser.RawConfigParser(allow_no_value=True)
        config.read_file(file)
    URL_1 = config.get("Settings", "URL")
    username = config.get("Settings", "Email")
    Pass = config.get("Settings", "Password")
    Vehicle = config.get("Settings", "Search-vehicle")
    #EngineParam1 = config.get("Settings", "Engine-parameters_1")




    #login process started
    driver.get(URL_1)
    driver.maximize_window()
    login_button = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@class='login_user mr-2']"))).click()
    time.sleep(3)
    Id = wait.until(
        EC.presence_of_element_located((By.XPATH, "//label[text()='Username']/following-sibling::input"))).send_keys(
        username)
    Password = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@id='password-field']"))).send_keys(Pass)
    SendOTP_button = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@class='login_btn']"))).click()
    time.sleep(5)
    print('Done')

    # fetch OPT from URL
    try:
        api_url = 'http://10.159.68.53:8080/jhswebapi/api/Login/getotp/ankithajakkula1993@gmail.com'  # Example API URL
        response = requests.get(api_url)
        if response.status_code == 200:
            data = response.json()
            OTP_login = data.get("data", {}).get("otp")
            print(f"OTP: {OTP_login}")
            time.sleep(5)
            Send_OTP = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@name='otp_Email']"))).send_keys(OTP_login)
        else:
            print("Failed to fetch data from the API.")
    except:
        print('unsuccessfull loop')

    VerifyOTP_button = wait.until(EC.presence_of_element_located((By.XPATH,
                                                                  "/html/body/app-root/app-login-page/app-modal[1]/div[2]/div[2]/div/div/div/div/div/div[1]/form/div[5]/button")))
    VerifyOTP_button.click()
    time.sleep(10)

    #timestamp generator for CSV (Naming convention)
    timestamp = datetime.datetime.now().strftime('%d%m%y_%H%M%S')
    csv_filename = f'data_{timestamp}.csv'

    #CSV file writer
    try:
        value_elements = driver.find_elements(By.XPATH,"//p[text()]//parent::div[@class='text-center live_widget ng-star-inserted']//child::p[2]")  # Replace with appropriate CSS selector
        values = [element.text for element in value_elements]
        with open(csv_filename, mode="w", newline="") as csvfile:
            csvfile.write('')
            csv_writer = csv.writer(csvfile)
            csv_writer.writerow(["Serial no","Parameter Name","Parameter value"])
        print('write success')
    except:
        print('Value cant be added')
    time.sleep(3)

    #Dropdown element selection
    try:
        print('loop trying')
        wait.until(EC.presence_of_element_located((By.XPATH, "//ng-select[@placeholder='Search Vehicle']"))).click()
        time.sleep(5)
        wait.until(EC.presence_of_element_located((By.XPATH,"//span[text()='"+Vehicle+"']"))).click()
        print('loop success')
        time.sleep(2)
    except:
        print('failed')

    # Track button click
    Track = wait.until(EC.presence_of_element_located((By.XPATH, "//button[text()='Track']"))).click()
    time.sleep(2)

# for loop 2
    for i in range(0,len(list)):
        Customise = wait.until(EC.presence_of_element_located((By.XPATH, "//a[text()='Customise ']"))).click()
        time.sleep(2)
        Reset_button = wait.until(
            EC.presence_of_element_located((By.XPATH, "//app-modal[5]/div[2]/div[3]/div/button[1]"))).click()
        time.sleep(2)
        Checkbox_selection_2 = wait.until(EC.presence_of_element_located((By.XPATH, "//label[text()=' "+list[i]+" ']/preceding-sibling::input"))).click()
        time.sleep(2)
        Submit_checkbox = wait.until(
            EC.presence_of_element_located((By.XPATH, "//app-modal[5]/div[2]/div[3]/div/button[3]"))).click()
        time.sleep(2)

        # append values into CSV
        try:
            value_elements = driver.find_elements(By.XPATH,"//p[text()]//parent::div[@class='text-center live_widget ng-star-inserted']//child::p[2]")
            values = [element.text for element in value_elements]
            #csv_filename = "portal_values.csv"
            with open(csv_filename, mode="a",newline='') as csvfile:
                csvfile.write('')
                csv_writer = csv.writer(csvfile)
                for value in values:
                    csv_writer.writerow([i+1,list[i],value])
            print([i+1,list[i],value])
        except WebDriverException as e:
            print(f"WebDriver Exception: {e}")
        i += 1
    print('CSV write completed')

    workbook = openpyxl.load_workbook(csv_filename)
    # Select the worksheet you want to work with
    worksheet = workbook.active  # Change to the desired worksheet if not the first one
    # Iterate through columns and autofit each one
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0,1,2].column)  # Get the column letter (e.g., 'A', 'B', 'C', ...)
        # Calculate the maximum length of data in the column
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        adjusted_width = (max_length + 2)  # Add a little extra width for padding

        # Set the column width
        worksheet.column_dimensions[column_letter] = ColumnDimension(worksheet, min=column[0].column,
                                                                     max=column[0].column,
                                                                     width=adjusted_width)
        workbook.save(csv_filename)
jiohumsufer_login()
